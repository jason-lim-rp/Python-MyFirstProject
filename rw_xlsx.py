from openpyxl import Workbook, load_workbook
wb = Workbook()

wb.create_sheet('data1')
print (wb.sheetnames)

# grab the active worksheet
ws = wb["data1"]

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
wb.close()


#Loading from a xlsx
wb2 = load_workbook("sample.xlsx")
ws2 = wb2["data1"]
c = ws2['A1']
print (c.value)



